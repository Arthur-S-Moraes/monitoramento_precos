from datetime import datetime
import openpyxl
from time import sleep
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as condicao_esperada
from selenium.common.exceptions import *



def monitoramento_precos(window, *sites):
    try:
        workbook = openpyxl.load_workbook('produtos.xlsx')
    except:
        workbook = openpyxl.Workbook()
        print('criando nova planilha')
        workbook['Sheet'].title = 'Kabum'
        workbook.create_sheet('mercadolivre')
        workbook.create_sheet('amazon')
        sheet_produto = workbook['Kabum']
        sheet_produto.append(['produto', 'Data Atual', 'preco', 'site'])
        sheet_produto = workbook['mercadolivre']
        sheet_produto.append(['produto', 'Data Atual', 'preco', 'site'])
        sheet_produto = workbook['amazon']
        sheet_produto.append(['produto', 'Data Atual', 'preco', 'site'])

    window.write_event_value('iniciando_automacao','iniciando o navegador')
    def iniciar_driver():
            print('iniciando o programa, aguarde alguns instantes')
            chrome_option = Options()
            arguments = ['--lang=en-US', '--start-maximized', '--force-dark-mode', '--log-level=3']
            for argument in arguments:
                chrome_option.add_argument(argument)


            driver = webdriver.Chrome(service=ChromeService(
                ChromeDriverManager().install()), options=chrome_option)
            
            wait = WebDriverWait(
                driver,
                10,
                poll_frequency=1,
                ignored_exceptions=[
                    NoSuchElementException,
                    ElementNotVisibleException,
                    ElementNotSelectableException
                ]
            )
            return driver, wait
    def selenium_planilha():
        driver, wait = iniciar_driver()
        urls = sites[0]
        for site in urls:
            if 'kabum' in site:
                driver.get(site)
                produto = wait.until(condicao_esperada.visibility_of_element_located((By.XPATH, "//h1[contains(@class,'sc-3a2f2afd-6 fWWWGi')]"))).text
                preco = wait.until(condicao_esperada.visibility_of_element_located((By.XPATH, "//h4[@class='sc-5492faee-2 ipHrwP finalPrice']"))).text
                sleep(2)
                sheet_produto = workbook['Kabum']
                print(f'dados do site Kabum salvos com sucesso')
            
            if 'mercadolivre' in site:
                driver.get(site)
                valor = wait.until(condicao_esperada.visibility_of_element_located((By.XPATH, "//span[@class='andes-money-amount ui-pdp-price__part andes-money-amount--cents-superscript andes-money-amount--compact']"))).text
                produto = wait.until(condicao_esperada.visibility_of_element_located((By.XPATH, "//h1[@class='ui-pdp-title']"))).text
                sleep(2)
                
                valor = valor.split('\n')
                try:
                    preco = valor[0] + valor[1] + valor[2] + valor[3]
                except:
                    preco = valor[0] + valor[1]
                sheet_produto = workbook['mercadolivre']
                print(f'dados do site Mercado Livre salvos com sucesso')
            
            if 'amazon' in site:
                driver.get(site)
                produto = wait.until(condicao_esperada.visibility_of_element_located((By.ID, "productTitle"))).text
                valor = wait.until(condicao_esperada.visibility_of_element_located((By.XPATH, "//span[@class='a-price aok-align-center reinventPricePriceToPayMargin priceToPay']"))).text
                sleep(2)

                valor = valor.split('\n')
                preco = valor[0] + ',' + valor[1]
                preco
                sheet_produto = workbook['amazon']
                print(f'dados do site da Amazon salvos com sucesso')
            
            data_atual = datetime.now()
            data_formatada = data_atual.strftime('%d/%m/%Y %H:%M')
            for linha in sheet_produto.iter_rows(min_row=2, max_row=1000,min_col=1, max_col=4):
                if linha[0].value == produto and linha[3].value == site:
                    linha[1].value = data_formatada
                    linha[2].value = preco
                    break
                elif linha[0].value == None:
                    linha[0].value = produto
                    linha[1].value = data_formatada
                    linha[2].value = preco
                    linha[3].value = site
                    break
                    
            workbook.save('produtos.xlsx')
            window.write_event_value('planilha_atualizada','Planilha salva com sucesso')
    selenium_planilha()